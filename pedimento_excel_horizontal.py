#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xml_to_excel_horizontal_all.py
Convierte cualquier XML en un Excel con tablas horizontales.
Incluye nodos anidados (por ejemplo <ProveedorComprador>) y concatena el nombre del XML en el Excel generado.
Uso:
    python3 xml_to_excel_horizontal_all.py archivo.xml
"""

import sys
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from textwrap import shorten

# =============================
# 🔹 Estilos
# =============================
border_thin = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000"),
)
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
bold_font = Font(bold=True)
wrap_alignment = Alignment(wrap_text=True, vertical="top")
center_alignment = Alignment(horizontal="center", vertical="center")

# ============================================================
# 🔹 Convierte un elemento y sus hijos en pares clave/valor
# ============================================================
def element_to_row(element):
    """Devuelve lista de (campo, valor) incluyendo vacíos."""
    row = []
    for child in element:
        tag = shorten(child.tag, width=30, placeholder="…")
        text = (child.text or "").strip()
        row.append((tag, text if text else "null"))
    return row

# ============================================================
# 🔹 Escribir bloque horizontal
# ============================================================
def write_block(ws, title, rows, start_row):
    """Escribe una tabla horizontal."""
    current_row = start_row
    ws.cell(row=current_row, column=1, value=title).font = bold_font
    ws.cell(row=current_row, column=1).fill = header_fill
    ws.cell(row=current_row, column=1).alignment = center_alignment
    current_row += 1

    col = 1
    max_per_line = 3
    for i, (campo, valor) in enumerate(rows, start=1):
        ws.cell(row=current_row, column=col, value=campo).font = bold_font
        ws.cell(row=current_row, column=col + 1, value=valor)
        ws.cell(row=current_row, column=col).border = border_thin
        ws.cell(row=current_row, column=col + 1).border = border_thin
        col += 2
        if i % max_per_line == 0:
            current_row += 1
            col = 1
    return current_row + 2

# ============================================================
# 🔹 Recursivo: procesa toda la estructura
# ============================================================
def process_node(ws, node, current_row, prefix=""):
    """Procesa recursivamente nodos anidados."""
    title = f"{prefix}{node.tag}"
    rows = element_to_row(node)
    if rows:
        current_row = write_block(ws, title, rows, current_row)

    for sub in node:
        current_row = process_node(ws, sub, current_row, prefix="↳ ")
    return current_row

# ============================================================
# 🔹 Principal
# ============================================================
def xml_to_excel_horizontal(xml_file):
    path = Path(xml_file)
    if not path.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {xml_file}")

    tree = ET.parse(path)
    root = tree.getroot()

    wb = Workbook()
    ws = wb.active
    ws.title = path.stem[:31]

    current_row = 1
    ws.cell(row=current_row, column=1, value=f"XML: {path.name}").font = Font(bold=True, size=14)
    current_row += 2

    # Procesar recursivamente todos los niveles
    current_row = process_node(ws, root, current_row)

    # Ajustar anchos de columnas
    for col in range(1, 15):
        ws.column_dimensions[chr(64 + col)].width = 25

    # ✅ Guardar Excel concatenando el nombre del XML
    output_file = f"{path.stem}_estructura_horizontal.xlsx"
    wb.save(output_file)
    print(f"✅ Archivo generado: {output_file}")

# ============================================================
# 🔹 Ejecución por línea de comandos
# ============================================================
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("❌ Uso: python3 xml_to_excel_horizontal_all.py 5002863.xml")
        sys.exit(1)

    xml_file = sys.argv[1]
    xml_to_excel_horizontal(xml_file)
