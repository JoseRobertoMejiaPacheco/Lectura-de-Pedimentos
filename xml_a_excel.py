#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pedimento_excel_full_extended.py
Convierte un pedimento XML en un Excel con:
 - Resumen
 - Facturas
 - Fracciones (productos)
 - ItemsDetallados (productos individuales dentro de fracciones)
 - Impuestos / Gastos (detalles + prorrateados)

Uso:
    python3 pedimento_excel_full_extended.py pedimento.xml

Resultado:
    pedimento_full_extended.xlsx
"""

import sys
import xml.etree.ElementTree as ET
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from PedimentoBuilder.builder import get


# ============================================================
# 🔹 UTILIDADES
# ============================================================
def find_text(element, path):
    """Devuelve texto limpio de una ruta dentro del XML."""
    el = element.find(path)
    return el.text.strip() if el is not None and el.text else ""


def aplicar_formato_excel(path_excel):
    """Aplica formato general a todas las hojas del Excel."""
    wb = load_workbook(path_excel)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    bold = Font(bold=True)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.freeze_panes = "A2"

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 22

        for cell in ws[1]:
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path_excel)


# ============================================================
# 🔹 EXTRACCIÓN DE DATOS
# ============================================================
def extraer_resumen(root):
    """Extrae encabezado general del pedimento."""
    return {
        "NumeroPedimento": find_text(root, "NumerodePedimentoCompleto"),
        "TipoOperacion": find_text(root, "TipoOperacion"),
        "ClavePedimento": find_text(root, "ClaveDePedimento"),
        "Regimen": find_text(root, "Regimen"),
        "ValorDolares": find_text(root, "ValorDolares"),
        "ValorAduana": find_text(root, "ValorAduana"),
        "TipoCambio": find_text(root, "TipoDeCambio"),
        "PesoBruto": find_text(root, "PesoBruto"),
        "FechaPago": find_text(root, "FechaDePagoDelPedimento"),
        "Cliente": find_text(root, "Cliente/RazonSocial"),
        "ClienteRFC": find_text(root, "Cliente/RFC"),
        "AgenteAduanal": find_text(root, "AgenteAduanal/Nombre"),
        "AgenteRFC": find_text(root, "AgenteAduanal/RFC"),
    }


def extraer_facturas(root):
    """Extrae todas las facturas del pedimento."""
    facturas = []
    for fac in root.findall("Facturas/Factura"):
        facturas.append({
            "Folio": find_text(fac, "Folio"),
            "Fecha": find_text(fac, "Fecha"),
            "Moneda": find_text(fac, "MonedaFactura"),
            "ValorDolares": find_text(fac, "ValorDolares"),
            "ValorTotal": find_text(fac, "ValorTotal"),
            "Proveedor": find_text(fac, "ProveedorComprador/RazonSocial"),
            "RFCProveedor": find_text(fac, "ProveedorComprador/RfcTaxId"),
            "PaisProveedor": find_text(fac, "ProveedorComprador/Pais"),
        })
    return pd.DataFrame(facturas)


def extraer_fracciones(root):
    fracs = []
    items_det = []

    for fr in root.iter("Fraccion"):
        valor_aduana = float(find_text(fr, "ValorAduana") or 0)
        valor_dolares = float(find_text(fr, "ValorDolares") or 0)
        num_frac = find_text(fr, "NumeroFraccion")
        descripcion = find_text(fr, "Descripcion")
        cantidad_fac = float(find_text(fr, "CantidadFactura") or 0)
        unidad_fac = find_text(fr, "UnidadFactura")

        if num_frac:
            fracs.append({
                "NumeroFraccion": num_frac,
                "Descripcion": descripcion,
                "CantidadFactura": cantidad_fac,
                "UnidadFactura": unidad_fac,
                "ValorAduana": valor_aduana,
                "ValorDolares": valor_dolares,
            })

        # Items dentro de fracción (en cualquier nivel)
        for it in fr.iter("Item"):
            if not get(it, "ItemNumber"):
                continue

            cantidad_it = float(find_text(it, "Cantidad") or 0)
            precio_it = float(find_text(it, "PrecioUnitario") or 0)
            total_it = cantidad_it * precio_it
            marca = find_text(it, "DescripcionesEspecificas/DescripcionEspecifica/Marca")
            modelo = find_text(it, "DescripcionesEspecificas/DescripcionEspecifica/Modelo")
            serie = find_text(it, "DescripcionesEspecificas/DescripcionEspecifica/Serie")
            datoid = find_text(it, "DescripcionesEspecificas/DescripcionEspecifica/DatoIdentificacion")

            items_det.append({
                "NumeroFraccion": num_frac,
                "Marca": marca,
                "Modelo": modelo,
                "Serie": serie,
                "DatoIdentificacion": datoid,
                "CantidadItem": cantidad_it,
                "PrecioUnitarioItem": precio_it,
                "TotalItem": total_it,
            })

    return pd.DataFrame(fracs), pd.DataFrame(items_det)


def extraer_impuestos_gastos(root):
    """Extrae impuestos y gastos incrementables."""
    regs = []

    # Impuestos por fracción
    for fr in root.findall("Fracciones/Fraccion"):
        num_frac = find_text(fr, "NumeroFraccion")
        for imp in fr.findall("Impuestos/Contribucion"):
            regs.append({
                "NumeroFraccion": num_frac,
                "Tipo": "IMPUESTO",
                "Clave": find_text(imp, "ClaveImpuesto"),
                "Concepto": find_text(imp, "ConceptoImpuesto"),
                "Importe": float(find_text(imp, "Importe") or 0),
            })

    # Gastos (incrementables)
    for pago in root.findall("Incrementables/OtrosPagos"):
        regs.append({
            "NumeroFraccion": "",
            "Tipo": "GASTO",
            "Clave": "GASTO",
            "Concepto": find_text(pago, "Concepto"),
            "Importe": float(find_text(pago, "ImporteMN") or 0),
        })

    return pd.DataFrame(regs)


def prorrateo(df_fracs, df_regs):
    """Distribuye gastos proporcionalmente al valor en aduana."""
    if df_regs.empty or df_fracs.empty:
        return df_fracs

    df = df_fracs.copy()
    total_val_aduana = df["ValorAduana"].sum() or 1
    df["GastoAsignado"] = 0.0

    for _, reg in df_regs[df_regs["Tipo"] == "GASTO"].iterrows():
        monto = reg["Importe"]
        df["GastoAsignado"] += (df["ValorAduana"] / total_val_aduana) * monto

    df["CostoTotal"] = df["ValorAduana"] + df["GastoAsignado"]
    df["CostoUnitario"] = df["CostoTotal"] / df["CantidadFactura"].replace(0, 1)
    return df


# ============================================================
# 🔹 PROGRAMA PRINCIPAL
# ============================================================
def main():
    input_file = sys.argv[1] if len(sys.argv) > 1 else "5004477.xml"
    path = Path(input_file)

    if not path.exists():
        print(f"❌ Archivo no encontrado: {input_file}")
        sys.exit(1)

    print(f"📄 Procesando archivo: {path.name}")
    tree = ET.parse(str(path))
    root = tree.getroot()

    # Extraer datos
    resumen = extraer_resumen(root)
    df_resumen = pd.DataFrame([resumen])
    df_facturas = extraer_facturas(root)
    df_fracs, df_items = extraer_fracciones(root)
    df_regs = extraer_impuestos_gastos(root)
    df_fracs_cost = prorrateo(df_fracs, df_regs)

    # Generar Excel
    output_xlsx = path.stem + "_full_extended.xlsx"
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        df_resumen.to_excel(writer, index=False, sheet_name="Resumen")
        df_facturas.to_excel(writer, index=False, sheet_name="Facturas")
        df_fracs_cost.to_excel(writer, index=False, sheet_name="Fracciones")
        df_items.to_excel(writer, index=False, sheet_name="ItemsDetallados")
        df_regs.to_excel(writer, index=False, sheet_name="ImpuestosGastos")

    aplicar_formato_excel(output_xlsx)
    print(f"✅ Archivo Excel generado correctamente: {output_xlsx}")


if __name__ == "__main__":
    main()
#python3 pedimento_excel_full_extended.py 5004477.xml
