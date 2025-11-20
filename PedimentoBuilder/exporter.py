import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


# ===========================================================
# 1) COSTOS POR ITEM (INCLUYE COSTO UNITARIO)
# ===========================================================
def df_costos_por_item(pedimento):
    rows = []

    for factura in pedimento.facturas:
        proveedor = factura.proveedor

        for fr in pedimento.fracciones:

            # -----------------------------------------
            # Contribuciones válidas
            # -----------------------------------------
            contribuciones_validas = [
                c for c in fr.contribuciones if float(c.tipo_de_tasa) != 0
            ]
            total_contribuciones = sum(float(c.importe) for c in contribuciones_validas)

            # -----------------------------------------
            # AGRUPAR ITEMS POR item_number
            # -----------------------------------------
            items_agrupados = {}

            for it in fr.items:
                key = it.item_number

                # Valor aduana REAL del item
                # Si viene vacío, usar total del item o fallback
                valor_aduana_item = float(it.total) if it.total > 0 else None

                if key not in items_agrupados:
                    items_agrupados[key] = {
                        "cantidad": 0.0,
                        "valor_aduana_item": valor_aduana_item,
                    }

                items_agrupados[key]["cantidad"] += float(it.cantidad)

                # Si el item trae valor aduana válido, conservarlo
                if valor_aduana_item is not None:
                    items_agrupados[key]["valor_aduana_item"] = valor_aduana_item

            # -----------------------------------------
            # PRORRATEO DTA
            # -----------------------------------------
            cantidad_total_fr = sum(float(it.cantidad) for it in fr.items)
            dta_total = float(fr.dta)
            dta_unitario = dta_total / cantidad_total_fr if cantidad_total_fr > 0 else 0

            # -----------------------------------------
            # CREAR UNA SOLA LÍNEA POR ITEM
            # -----------------------------------------
            for item_number, data in items_agrupados.items():

                cantidad_item = data["cantidad"]

                # valor_aduana del item (si existe)
                if data["valor_aduana_item"] is not None:
                    valor_aduana = data["valor_aduana_item"]
                else:
                    # si no existe, prorratear el valor_aduana de la fracción
                    valor_aduana = float(fr.valor_aduana) / cantidad_total_fr * cantidad_item

                # calcular DTA proporcional
                dta_item = dta_unitario * cantidad_item

                # costo total real del item
                costo_total_item = valor_aduana + total_contribuciones + dta_item

                costo_unitario = costo_total_item / cantidad_item if cantidad_item > 0 else 0

                rows.append({
                    "item_number": item_number,
                    "cantidad": cantidad_item,
                    "valor_aduana": valor_aduana,
                    "contribuciones_validas": total_contribuciones,
                    "dta_prorrateado": dta_item,
                    "costo_total_item": costo_total_item,
                    "costo_unitario": costo_unitario,
                    "proveedor": proveedor.razon_social
                })

    return pd.DataFrame(rows)

# ===========================================================
# 2) CONTRIBUCIONES DETALLADAS (INCLUYE item_number)
# ===========================================================
def df_contribuciones_detalle(pedimento):
    rows = []

    for fr in pedimento.fracciones:

        contribs_validas = [
            c for c in fr.contribuciones if float(c.tipo_de_tasa) != 0
        ]

        for it in fr.items:

            for c in contribs_validas:
                rows.append({
                    "item_number": it.item_number,
                    "cantidad": float(it.cantidad),
                    "total_item": float(it.total),

                    "valor_aduana_fraccion": float(fr.valor_aduana),

                    "concepto_impuesto": c.concepto_impuesto,
                    "importe_contribucion": float(c.importe),
                    "tasa": float(c.tipo_de_tasa),
                })

    return pd.DataFrame(rows)


# ===========================================================
# 3) EXPORTAR EXCEL PREMIUM (SIN FORMATO MONEDA)
# ===========================================================
def exportar_excel_pedimento_premium(
        pedimento, df_costos, df_contrib, output_path, logo_path=None
    ):

    wb = Workbook()

    # ------------------------------------
    # ESTILOS
    # ------------------------------------
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=18)
    label_font = Font(bold=True, size=14)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # ===========================================================
    # HOJA 1 - PORTADA
    # ===========================================================
    ws = wb.active
    ws.title = "Portada"

    ws.merge_cells("A1:D1")
    ws["A1"] = "REPORTE DE COSTOS POR PEDIMENTO"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Número de Pedimento:"
    ws["B3"] = pedimento.numero

    ws["A4"] = "Fecha Pago:"
    ws["B4"] = pedimento.fecha_pago

    ws["A5"] = "Proveedor:"
    ws["B5"] = pedimento.facturas[0].proveedor.razon_social

    ws["A3"].font = ws["A4"].font = ws["A5"].font = label_font

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 35

    # ===========================================================
    # HOJA 2 - COSTOS POR ITEM
    # ===========================================================
    ws2 = wb.create_sheet("Costos por Item")

    # Encabezados
    for col_num, col_name in enumerate(df_costos.columns, 1):
        cell = ws2.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.border = thin_border
        cell.fill = gray_fill
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for row_num, row in enumerate(df_costos.itertuples(index=False), 2):
        for col_num, value in enumerate(row, 1):
            cell = ws2.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            # NOTA: SIN FORMATO DE MONEDA

    # Autoajuste columnas
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = max(len(str(c.value)) for c in col) + 2

    # -----------------------------
    # TOTAL GENERAL
    # -----------------------------
    last_row = len(df_costos) + 2

    ws2[f"A{last_row}"] = "TOTAL GENERAL:"
    ws2[f"A{last_row}"].font = Font(bold=True)

    # Columna 'costo_total_item' = E
    ws2[f"E{last_row}"] = f"=SUM(E2:E{last_row-1})"
    ws2[f"E{last_row}"].font = Font(bold=True)
    ws2[f"E{last_row}"].border = thin_border
    # (SIN FORMATO DE MONEDA)

    # ===========================================================
    # HOJA 3 - CONTRIBUCIONES DETALLE
    # ===========================================================
    ws3 = wb.create_sheet("Contribuciones")

    # Encabezados
    for col_num, col_name in enumerate(df_contrib.columns, 1):
        cell = ws3.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.border = thin_border
        cell.fill = gray_fill
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for row_num, row in enumerate(df_contrib.itertuples(index=False), 2):
        for col_num, value in enumerate(row, 1):
            cell = ws3.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            # SIN FORMATO DE MONEDA

    # Autoajuste columnas
    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = max(len(str(c.value)) for c in col) + 2

    # ===========================================================
    # GUARDAR ARCHIVO
    # ===========================================================
    wb.save(output_path)
